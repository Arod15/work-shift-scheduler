import openpyxl as op
import datetime
from openpyxl.styles import PatternFill

days = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

filename = 'init.xlsx'
# filename = 'new-init.xlsx'
workbook = op.load_workbook(filename)
sheet1 = workbook.active

def print_sheet():
    maxr = sheet1.max_row
    maxc = sheet1.max_column
    for i in range(1, maxr+1):
        for j in range(1, maxc+1):
            print(i, j, sheet1.cell(row=i, column=j).value)

def color_row(r):
  for i in range(1, sheet1.max_column+1):
    sheet1.cell(row=r, column=i).fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
def insert_hours(start, end, date):
#   if date > datetime.date.today():
#       raise ValueError('Can\'t insert into a date that hasn\'t occured yet.')
  i = 1
  j = 2
  while sheet1.cell(row=i, column=2).value != date:
    # print(sheet1.cell(row=i, column=2).value, date)
    print(i, j, sheet1.cell(row=i, column=j).value, date)
    i += 1
  while sheet1.cell(row=i, column=j).value is not None:
    print(i, j, sheet1.cell(row=i, column=j).value)
    j += 1
  sheet1.cell(row=i, column=j).value = start
  print(f"inserted {start} at ({i}, {j})")
  sheet1.cell(row=i, column=j+1).value = end
  print(f"inserted {end} at ({i}, {j})")
def sum_hours(r):
  total = 0
  for i in range(1, 8):
    total += sheet1.cell(row=r+i, column=7).value
  sheet1.cell(row=r, column=8).value = total
def calc_hours(r):
  array = []
  for i in range(3, sheet1.max_column, 2):
    start = sheet1.cell(row=r, column=i).value
    end = sheet1.cell(row=r, column=i+1).value
    if start is None or end is None:
      array.append(0)
    else:
      converted_start = datetime.datetime.combine(datetime.date.today(), start)
      converted_end = datetime.datetime.combine(datetime.date.today(), end)
      diff = converted_start - converted_end
      shift_hours = abs(diff.total_seconds() / 3600)
      array.append(shift_hours)
  daily_hours = sum(array)
  sheet1.cell(row=r, column=sheet1.max_column).value = daily_hours
def shifts(row, num):
  col = 0
  for i in range(num):
    sheet1.cell(row=row, column=(3+(num*i))).value = 'Start'
    sheet1.cell(row=row, column=(4+(num*i))).value = 'End'
    if i == num-1:
      col = 5 + (num * i)
  sheet1.cell(row=row, column=col).value = 'Hours'
def week(row):
  for i in range(len(days)):
    sheet1.cell(row=row+i, column=1).value = days[i]
def dates(row, col):
  start_date = sheet1.cell(row=row, column=col).value
  res = start_date
  for i in range(1, len(days)):
    res += datetime.timedelta(days=1)
    # cons_start_date = datetime.date(res.year, res.month, res.day)
    sheet1.cell(row=row+i, column=col).value = res
    # start_date += datetime.timedelta(days=1)
def generate_one(r, c, start):
  sheet1.cell(row=r, column=1).value = 'Day'
  sheet1.cell(row=r, column=2).value = 'Date'
  shifts(r, 2)
  week(r+1)
  sheet1.cell(row=r+1, column=2).value = start
  dates(r+1, 2)
  color_row(r+8)
# def main():
#   i = 2
#   day_date = datetime.date.today()
#   while sheet1.cell(row=i, column=2).value is not None:
#     day_date = sheet1.cell(row=i, column=2).value
#     sheet1.cell(row=i, column=2).value = datetime.date(day_date.year, day_date.month, day_date.day)
#     i += 1
#   i += 1
#   start_date = day_date + datetime.timedelta(days=1)
#   new_day_date = datetime.date(start_date.year, start_date.month, start_date.day)
#   init(i, 1, start=new_day_date)
zero_time = datetime.time(hour=0, minute=0, second=0)
begin = datetime.date(2019, 5, 12)
date = datetime.datetime.combine(begin, zero_time)

if sheet1.cell(row=1, column=1).value is None:
    for x in range(4):
        generate_one(1 + 9*x, 1, date)
        date += datetime.timedelta(days=7)
# if sheet1.cell(row=1, column=1).value is None:
#   generate_one(1, 1)
# else:
#   main()
#   print('mained')


# start = datetime.time(hour=5, minute=30)
# end = datetime.time(hour=9, minute=45)
# now = datetime.date.today() - datetime.timedelta(days=3)
# print('now: ', now)
# time_now = datetime.datetime.combine(now, zero_time)
# insert_hours(start, end, time_now)

# start = datetime.time(hour=5, minute=30)
# end = datetime.time(hour=9, minute=45)
# now = datetime.date.today() + datetime.timedelta(days=2)
# print('now: ', now)
# insert_hours(start, end, now)

# calc_hours(7)
# calc_hours(11)

# print_sheet()

workbook.save('new-' + filename)
# workbook.save(filename)
workbook.close()
