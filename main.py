import openpyxl as op
import datetime
from openpyxl.styles import PatternFill

days = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

filename = 'sample.xlsx'
# filename = 'NEW-sample.xlsx'
workbook = op.load_workbook(filename)
sheet1 = workbook.active

def print_sheet(maxr, maxc):
  for i in range(1, maxr+1):
    for j in range(1, maxc+1):
      print(i, j, sheet1.cell(row=i, column=j).value)
def color_row(r):
  for i in range(1, sheet1.max_column+1):
    sheet1.cell(row=r, column=i).fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
def insert_hours(start, end, date):
  i = 1
  # while i < 10:
  #   print(sheet1.cell(row=i, column=2).value)
  #   i += 1
  while sheet1.cell(row=i, column=2).value is not None and sheet1.cell(row=i, column=2).value != date:
    print(sheet1.cell(row=i, column=2).value, date)
    i += 1
  j = 1
  while sheet1.cell(row=i, column=j).value is not None:
    j += 1
  sheet1.cell(row=i, column=j).value = start
  sheet1.cell(row=i, column=j+1).value = end
def sum_hours(r):
  total = 0
  for i in range(1, 8):
    total += sheet1.cell(row=r+i, column=7).value
  sheet1.cell(row=r, column=8).value = total
def calc_hours(r):
  array = []
  for i in range(3, sheet1.max_column, 2):
    start = sheet1.cell(row=r, column=i).value
    end = sheet1.cell(row=r, column=i+1).value
    if start is None or end is None:
      array.append(0)
    else:
      converted_start = datetime.datetime.combine(datetime.date.today(), start)
      converted_end = datetime.datetime.combine(datetime.date.today(), end)
      diff = converted_start - converted_end
      shift_hours = abs(diff.total_seconds() / 3600)
      array.append(shift_hours)
  daily_hours = sum(array)
  sheet1.cell(row=r, column=sheet1.max_column).value = daily_hours
def shifts(row, num):
  col = 0
  for i in range(num):
    sheet1.cell(row=row, column=(3+(num*i))).value = 'Start'
    sheet1.cell(row=row, column=(4+(num*i))).value = 'End'
    if i == num-1:
      col = 5 + (num * i)
  sheet1.cell(row=row, column=col).value = 'Hours'
def week(row):
  for i in range(len(days)):
    sheet1.cell(row=row+i, column=1).value = days[i]
def dates(row, col):
  start_date = sheet1.cell(row=row, column=col).value
  for i in range(1, len(days)):
    start_date += datetime.timedelta(days=1)
    sheet1.cell(row=row+i, column=col).value = start_date
def init(r, c, start=None):
  sheet1.cell(row=r, column=1).value = 'Day'
  sheet1.cell(row=r, column=2).value = 'Date'
  shifts(r, 2)
  week(r+1)
  if start is None:
    sheet1.cell(row=r+1, column=2).value = datetime.date(datetime.date.today().year, 5, 26)
  else:
    sheet1.cell(row=r+1, column=2).value = start
  dates(r+1, 2)
  color_row(r+8)
def main():
  i = 1
  day_date = datetime.date.today()
  while sheet1.cell(row=i, column=1).value is not None:
    day_date = sheet1.cell(row=i, column=2).value
    i += 1
  i += 1
  start_date = day_date + datetime.timedelta(days=1)
  init(i, 1, start=start_date)

if sheet1.cell(row=1, column=1).value is None:
  init(1, 1)
  print('inited')
else:
  main()
  print('mained')
start = datetime.time(hour=5, minute=30)
end = datetime.time(hour=9, minute=45)
now = datetime.date.today()
insert_hours(start, end, now)

calc_hours(7)

print_sheet(sheet1.max_row, sheet1.max_column)

# workbook.save('new-' + filename)
workbook.save(filename)
workbook.close()
